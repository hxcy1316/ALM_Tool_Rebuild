import openpyxl
import os
from functools import reduce
from win32com.client import Dispatch


class EXCEL():
    def __init__(self, result_file_path, test_instance_list):
        self.result_file_path = result_file_path
        self.test_instance_list = test_instance_list
        self.unique_instance_dict = {}
        self.wb = openpyxl.Workbook()
        self.sheet_execution_detail = "Execution Details"
        self.sheet_unique_execution_detail = "Unique Execution Details"
        self.sheet_execution_summary = "Execution Summary"

    def open_sheet(self, sheet_name):
        if not os.path.exists(self.result_file_path):
            print("File Doesn't Exist")
            self.wb.save(self.result_file_path)
            self.wb.close()
        self.wb = openpyxl.load_workbook(self.result_file_path)
        print("Load workbook {} successfully".format(self.result_file_path))
        if sheet_name in self.wb.sheetnames:
            print("Sheet already exist, remove and recreate the sheet {}".format(sheet_name))
            self.wb.remove(self.wb[sheet_name])
        self.wb.create_sheet(sheet_name)
        return self.wb[sheet_name]

    def get_execution_detail(self):
        sheet_column_names = [
            "Test ID",
            "L1 Feature",
            "L2 Feature",
            "L3 Feature",
            "L4 Feature",
            "Test_instance_path",
            "Test_set_name",
            "Test Instance ID",
            "Test_instance_status"
        ]
        work_sheet = self.__initial_column_name(self.sheet_execution_detail, sheet_column_names)
        print("Start initial table in {}".format(work_sheet))
        row = 2
        print("Start filling the data ...")
        for test_instance in self.test_instance_list:
            work_sheet.cell(row, 1).value = test_instance["test_instance_test_id"]
            work_sheet.cell(row, 2).value = test_instance["test_instance_L1"]
            work_sheet.cell(row, 3).value = test_instance["test_instance_L2"]
            work_sheet.cell(row, 4).value = test_instance["test_instance_L3"]
            work_sheet.cell(row, 5).value = test_instance["test_instance_L4"]
            work_sheet.cell(row, 6).value = test_instance["test_instance_path"]
            work_sheet.cell(row, 7).value = test_instance["test_set_name"]
            work_sheet.cell(row, 8).value = test_instance["test_instance_id"]
            work_sheet.cell(row, 9).value = test_instance["test_instance_status"]
            row = row + 1
        print("End filling data in sheet {}".format(self.sheet_execution_detail))
        self.wb.save(self.result_file_path)

    def get_unique_execution_detail(self):
        sheet_column_names = [
            "Test ID",
            "L1 Feature",
            "L2 Feature",
            "L3 Feature",
            "L4 Feature",
            "Unique Status"
        ]
        work_sheet = self.__initial_column_name(self.sheet_unique_execution_detail, sheet_column_names)
        print("Start initial table in {}".format(work_sheet))
        row = 2
        # unique_instance_dict = {<id>: <status>, <id>: <status>, ...}
        self.unique_instance_dict = self.__format_to_unique_id_status_list(self.__format_to_id_status_list(self.test_instance_list))
        # print(self.unique_instance_dict)
        print("Start filling the data ...")
        for instance_test_id, instance_test_status in self.unique_instance_dict.items():
            test_id = instance_test_id
            test_status = instance_test_status
            test_property_dict = self.__get_properties_by_id(test_id, self.test_instance_list)
            work_sheet.cell(row, 1).value = test_id
            work_sheet.cell(row, 2).value = test_property_dict["test_instance_L1"]
            work_sheet.cell(row, 3).value = test_property_dict["test_instance_L2"]
            work_sheet.cell(row, 4).value = test_property_dict["test_instance_L3"]
            work_sheet.cell(row, 5).value = test_property_dict["test_instance_L4"]
            work_sheet.cell(row, 6).value = test_status
            row = row + 1
        print("End filling data in sheet {}".format(self.sheet_unique_execution_detail))
        self.wb.save(self.result_file_path)

    def load_array_to_sheet(self, arr, sheet_name):
        sheet_column_names = [
            'L1 Feature',
            'L2 Feature',
            'L3 Feature',
            'L4 Feature',
            'Unique_Planned_Test',
            'Unique_Executed_Test',
            'Unique_Passed_Test',
            'Unique_Execution_Rate',
            'Unique_Pass_Rate',
            'Total_Planed_Test',
            'Total_Executed_Test',
            'Total_Passed_Test',
            'Total_Execution_Rate',
            'Total_Pass_Rate'
        ]
        sheet = self.__initial_column_name(sheet_name, sheet_column_names)
        for row in arr:
            sheet.append(row)
        self.wb.save(self.result_file_path)
        return sheet

    def format_execution_summary(self, sheet):
        ws_summary = sheet
        row_count = ws_summary.max_row
        for row in ws_summary.iter_rows(min_row=2, max_row=row_count, min_col=5, max_col=14):
            # row[0 - 4]: Unique [Planned, Executed, Passed, Execution Rate, Pass Rate]
            # row[5 - 9]: Total [Planned, Executed, Passed, Execution Rate, Pass Rate]
            row[3].number_format = '0%'
            row[4].number_format = '0%'
            row[8].number_format = '0%'
            row[9].number_format = '0%'
        self.wb.save(file_path)
        return ws_summary

    def caculate_percentage(self, up_value, down_value):
        if down_value == 0:
            percentage = 0
        else:
            percentage = round(up_value / down_value, 2)
        return percentage

    def __initial_column_name(self, sheet_name, name_list):
        sheet = self.open_sheet(sheet_name)
        sheet.append(name_list)
        self.wb.save(self.result_file_path)
        return sheet

    def __get_properties_by_id(self, test_id, instance_list):
        for instance in instance_list:
            if test_id in instance.values():
                return instance

    def __format_to_id_status_list(self, full_instance_list):
        # id_status_list = []
        # for instance_dict in full_instance_list:
        #     id_status_list.append(self.__format_dict(instance_dict))
        id_status_list = map(self.__format_dict, full_instance_list)
        return id_status_list

    def __format_dict(self, temp_dict):
        formated_dict = {temp_dict["test_instance_test_id"]: temp_dict["test_instance_status"]}
        # print("formated dict is {}".format(formated_dict))
        return formated_dict

    def __format_to_unique_id_status_list(self, id_status_list):
        unique_id_status_list = reduce(self.__update_dict, id_status_list)
        return unique_id_status_list

    def __update_dict(self, dict1, dict2):
        for key, value in dict2.items():
            if key in dict1.keys():
                unique_status = self.__get_unique_status([value, dict1[key]])
                dict1.update({f'{key}': unique_status})
            else:
                dict1.update({f'{key}': dict2[key]})
        # print(dict1)
        return dict1

    def __get_unique_status(self, status_list):
        if "Failed" in status_list:
            return "Failed"
        elif "Passed" in status_list:
            return "Passed"
        else:
            return "No Run"

    def close(self):
        self.wb.close()


def format_execution_summary(file_path, sheet_name):
    try:
        excel_win32 = Dispatch("Excel.Application")
        excel_win32.Visible = False
        wb = excel_win32.Workbooks.Open(file_path)
        ws = wb.Worksheets(sheet_name)
        ws.Activate()
        myrange_NA = ws.Range(ws.cells(1, 1), ws.cells(ws.usedRange.rows.count, 4))
        for cell in myrange_NA:
            if cell.value == "NA":
                cell.value = ""
        wb.Save()
        myrange1 = ws.Range(ws.cells(1, 1), ws.cells(ws.usedRange.rows.count, 14))
        myrange1.Subtotal(
            GroupBy=1,
            Function=-4157,
            TotalList=[5, 6, 7, 8, 9, 10, 11, 12, 13, 14],
            Replace=False,
            PageBreaks=False,
            SummaryBelowData=True
        )
        myrange2 = ws.Range(ws.cells(1, 2), ws.cells(ws.usedRange.rows.count, 14))
        myrange2.Subtotal(
            GroupBy=1,
            Function=-4157,
            TotalList=[4, 5, 6, 7, 8, 9, 10, 11, 12, 13],
            Replace=False,
            PageBreaks=False,
            SummaryBelowData=True
        )
        myrange3 = ws.Range(ws.cells(1, 3), ws.cells(ws.usedRange.rows.count, 14))
        myrange3.Subtotal(
            GroupBy=1,
            Function=-4157,
            TotalList=[3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
            Replace=False,
            PageBreaks=False,
            SummaryBelowData=True
        )
        ws.Columns.AutoFit()
        wb.Save()
        for i in range(2, ws.usedRange.rows.count + 1):
            find_cell_list = [
                str(ws.cells(i, 1).value).find("Total", -5),
                str(ws.cells(i, 2).value).find("Total", -5),
                str(ws.cells(i, 3).value).find("Total", -5)
            ]
            if max(find_cell_list) >= 0:
                if ws.cells(i, 5).value == 0:
                    ws.cells(i, 8).value = 0
                    ws.cells(i, 9).value = 0
                else:
                    ws.cells(i, 8).value = round(ws.cells(i, 6).value / ws.cells(i, 5).value, 2)
                    ws.cells(i, 9).value = round(ws.cells(i, 7).value / ws.cells(i, 5).value, 2)
                if ws.cells(i, 10).value == 0:
                    ws.cells(i, 13).value = 0
                    ws.cells(i, 14).value = 0
                else:
                    ws.cells(i, 13).value = round(ws.cells(i, 11).value / ws.cells(i, 10).value, 2)
                    ws.cells(i, 14).value = round(ws.cells(i, 12).value / ws.cells(i, 10).value, 2)
        ws.Range(ws.cells(2, 8), ws.cells(ws.usedRange.rows.count, 9)).Style = "Percent"
        ws.Range(ws.cells(2, 13), ws.cells(ws.usedRange.rows.count, 14)).Style = "Percent"
        wb.Save()
        ws.cells.EntireColumn.Hidden = False
        ws.Range("A1:P1").WrapText = True
        ws.Range("A1:P1").Font.Bold = True
        ws.Range(ws.cells(1, 1), ws.cells(ws.usedRange.rows.count, ws.usedRange.columns.count)).Borders.LineStyle = 1
        ws.Range(ws.cells(1, 5), ws.cells(ws.usedRange.rows.count, 9)).Interior.Color = 220 + 230 * 256 + 241 * 256 * 256
        ws.Range(ws.cells(1, 10), ws.cells(ws.usedRange.rows.count, 14)).Interior.Color = 242 + 220 * 256 + 219 * 256 * 256
        ws.Range(ws.cells(1, 6), ws.cells(ws.usedRange.rows.count, 7)).EntireColumn.Hidden = True
        ws.Range(ws.cells(1, 11), ws.cells(ws.usedRange.rows.count, 12)).EntireColumn.Hidden = True
        ws.Range(ws.cells(1, 9), ws.cells(ws.usedRange.rows.count, 9)).Font.Bold = True
        ws.Range(ws.cells(1, 14), ws.cells(ws.usedRange.rows.count, 14)).Font.Bold = True
        wb.Save()
        print("Format finish, refer to '{}' for details".format(sheet_name))
    except Exception as e:
        print(e)
    finally:
        wb.close()
        wb = None
        excel_win32 = None


if __name__ == "__main__":
    pass
    # full_instance_list = ['1', '2']
    # file_path = r"C:\My Doc\My Github\ALM_Tool_REBUILD\test.xlsx"
    # # get_execution_detail(file_path, full_instance_list)
